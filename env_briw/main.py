import pandas as pd
import numpy as np
import os
import logging
import openpyxl
from env import InventoryEnvGYConfig
import train_ppo
import train_BS
import train_sQ
import train_dqn
import eval_ppo
import eval_BS
import eval_sQ
import eval_dqn
import plot
import BS
import sQ
import time
import pickle
from stable_baselines3 import PPO
from stable_baselines3 import DQN
import scipy.optimize as optimize
from scipy.optimize import Bounds
from scipy.optimize import show_options
from openpyxl import load_workbook
import json

# Set the following flags to True to run the corresponding algorithm

import scipy.optimize as optimize
import numpy as np

ppo = False
dqn = True
bs = False
sq = False

train = True
plot_train = True
eval = True
plot_eval =True

powell =False
brute_force_scipy = False
brute_force_manual = False

def main():
    # Load configurations from an Excel file
    excel_path = r'../rewardshaping/configurations.xlsx'
    config_json_path = r'../rewardshaping/config.json'
    df_configurations = pd.read_excel(excel_path, engine='openpyxl')
    configurations = df_configurations.to_dict('records')
    workbook = openpyxl.load_workbook(excel_path)
    sheet = workbook.active
        # Leggere il file di configurazione JSON
    with open(config_json_path, 'r') as file:
        json_config = json.load(file)

    # Adjust policy_kwargs and learning rate if needed
    learning_rate = 0.0003
    steps = list(range(1, 11))
    for i, config in enumerate(configurations):
        config_details = "_".join([f"{k}_{v}" for k, v in config.items() if k != 'configuration'])
        output_dir = f'{config_details}'
        env = InventoryEnvGYConfig(config,json_config)
        total_timesteps = 150000
        if ppo:

            n_steps = 2048
            batch_size = 128
            n_epochs = 10

            subdir = 'ppo'
            subdir_1 = f'lr_{learning_rate}_n_steps_{n_steps}_batch_size_{batch_size}_n_epochs_{n_epochs}'
            full_path = os.path.join(output_dir, subdir, subdir_1)
            os.makedirs(full_path, exist_ok=True)
            # Train and evaluate the model
            if train:
                model = train_ppo.train_ppo_model(env, learning_rate, total_timesteps, full_path, n_steps, batch_size, n_epochs)
            if plot_train:
                train_ppo.save_logs(full_path)
                plot.plot_reward_convergence(full_path)

            if eval:
                model_path = os.path.join(full_path, f"./logs/ppo_model")

                model = PPO.load(model_path)
                eval_ppo.evaluate_policy_and_log_detailed_metrics(model, env, full_path, n_eval_episodes=5)

            if plot_eval:
                eval_ppo.save_metrics_to_dataframe(full_path, config_details)
                plot.plot_episodes_metrics(steps, full_path)

        elif dqn:

            batch_size = 64
            buffer_size = 1000
            gradient_steps = 1
            target_update_interval = 1000

            subdir = 'dqn'
            subdir_1 = f'lr_{learning_rate}_batch_size_{batch_size}_buffer_size_{buffer_size}_gradient_steps_{gradient_steps}_target_update_interval_{target_update_interval}'
            full_path = os.path.join(output_dir, subdir, subdir_1)
            os.makedirs(full_path, exist_ok=True)

            if train:
                model = train_dqn.train_dqn_model(env, learning_rate, total_timesteps, full_path, batch_size, buffer_size, gradient_steps, target_update_interval)
            if plot_train:
                train_dqn.save_logs(full_path)
                plot.plot_reward_convergence(full_path)

            if eval:
                model_path = os.path.join(full_path, f"./logs/dqn_model")

                model = DQN.load(model_path)
                eval_dqn.evaluate_policy_and_log_detailed_metrics(model, env, full_path, n_eval_episodes=5)

            if plot_eval:
                eval_dqn.save_metrics_to_dataframe(full_path, config_details)
                plot.plot_episodes_metrics(steps, full_path) 

        elif bs:
            subdir = 'bs'
            full_path = os.path.join(output_dir, subdir)
            os.makedirs(full_path, exist_ok=True)
            if powell:
                bs_path = os.path.join(full_path, 'powell')
            if brute_force_scipy:
                bs_path = os.path.join(full_path, 'brute_force_scipy')
            if brute_force_manual:
                bs_path = os.path.join(full_path, 'brute_force_manual')
            os.makedirs(bs_path, exist_ok=True)
            
            base_stock_level_column = None
            if train:
                start_time = time.time()

                if powell:
                    np.random.seed(42)
                    bnds = [(5, 25)]
                    initial_guess = 5
                    res = optimize.minimize(train_BS.fun, initial_guess, args=(
                        env,), method='Powell', bounds=bnds, tol=0.00001)
                    
                    print("res:", np.around(res.x))
                    print("Best s:", np.around(res.x[0]))
                    print("Best average reward:", -res.fun)
            
                    full_path = os.path.join(bs_path, 'pickle_file')
                    os.makedirs(full_path, exist_ok=True)
                    filename = os.path.join(full_path, f'best_base_stock.pkl')
                    with open(filename, 'wb') as f:
                            pickle.dump(res.x[0], f)

                    for cell in sheet[1]:  # Assumendo che i titoli siano nella prima riga
                        if cell.value == 'base_stock_level':
                            base_stock_level_column = cell.column
                            break

                    if base_stock_level_column is None:
                        # Creare la colonna base_stock_level se non esiste
                        base_stock_level_column = sheet.max_column + 1
                        sheet.cell(row=1, column=base_stock_level_column, value='base_stock_level')
                
                    
                    sheet.cell(row=i+2, column=base_stock_level_column, value=np.around(res.x[0]))

                    # Salva il file
                    workbook.save(excel_path)


                if brute_force_scipy:
                    rranges = [slice(5, 15, 1)]
                    resbrute = optimize.brute(train_BS.fun, rranges, args=(
                        env,), full_output=True, finish=None)

                    end_time = time.time()
                    elapsed_time = end_time - start_time
                    print(elapsed_time)
                    print(resbrute[0])
                    print(resbrute[1])
                    print(resbrute[3])
                    grid_values = resbrute[3]
                    print(grid_values.size)

                    full_path = os.path.join(bs_path, 'pickle_file')
                    os.makedirs(full_path, exist_ok=True)
                    filename = os.path.join(full_path, f'levels.pkl')
                    with open(filename, 'wb') as f:
                        pickle.dump(list(range(5, 15)), f)
                    filename = os.path.join(full_path, f'avg_rewards.pkl')
                    with open(filename, 'wb') as f:
                            pickle.dump(-resbrute[3], f)
                    filename = os.path.join(full_path, f'best_base_stock.pkl')
                    with open(filename, 'wb') as f:
                            pickle.dump(resbrute[0], f)
                    
                    for cell in sheet[1]:  # Assumendo che i titoli siano nella prima riga
                        if cell.value == 'base_stock_level':
                            base_stock_level_column = cell.column
                            break

                    if base_stock_level_column is None:
                        # Creare la colonna base_stock_level se non esiste
                        base_stock_level_column = sheet.max_column + 1
                        sheet.cell(row=1, column=base_stock_level_column, value='base_stock_level')
                
                    
                    sheet.cell(row=i+2, column=base_stock_level_column, value=np.around(resbrute[0]))

                    # Salva il file
                    workbook.save(excel_path)
                
                if brute_force_manual:
                    base_stock_level = train_BS.train_bs_policy(env, bs_path, 5, 25, total_timesteps)

                    for cell in sheet[1]:  # Assumendo che i titoli siano nella prima riga
                        if cell.value == 'base_stock_level':
                            base_stock_level_column = cell.column
                            break

                    if base_stock_level_column is None:
                        # Creare la colonna base_stock_level se non esiste
                        base_stock_level_column = sheet.max_column + 1
                        sheet.cell(row=1, column=base_stock_level_column, value='base_stock_level')
                
                    
                    sheet.cell(row=i+2, column=base_stock_level_column, value=base_stock_level)
                    workbook.save(excel_path)

                end_time = time.time()

                elapsed_time = end_time - start_time
                print(elapsed_time)


            if plot_train:
                plot.plot_rewards_per_bs_level(bs_path)
            
            if eval:
                best_bs_path = os.path.join(bs_path, 'pickle_file/best_base_stock.pkl')
                with open(best_bs_path, 'rb') as file:
                    base_stock_level = pickle.load(file)
                bs_class = BS.BSpolicy(base_stock_level)
                eval_BS.evaluate_policy_and_log_detailed_metrics(env, bs_class, bs_path, n_eval_episodes=20)
            if plot_eval:
                eval_BS.save_metrics_to_dataframe(bs_path, config_details)
                plot.plot_episodes_metrics(steps, bs_path)

        elif sq:

            subdir = 'sq'
            full_path = os.path.join(output_dir, subdir)
            os.makedirs(full_path, exist_ok=True)

            if powell:
                sq_path = os.path.join(full_path, 'powell')
            #sq_path = os.path.join(full_path, 'brute_force_scipy')
            if brute_force_manual:
                sq_path = os.path.join(full_path, 'brute_force_manual')
            os.makedirs(sq_path, exist_ok=True)

            s_column = None
            Q_column = None
            if train:

               
                if powell:
                    bnds = [(0, 10), (0, 10)]
                    # Eseguiamo l'ottimizzazione
                    # Cambia la stima iniziale in base al tuo problema
                    initial_guess = (5, 5)
                    res = optimize.minimize(train_sQ.fun, initial_guess, args=(
                        env,), method='Powell', bounds=bnds, tol=0.000001, options={'disp': True})

                    print("Best s:", np.around(res.x[0]))
                    print("Best Q:", np.around(res.x[1]))
                    print("Best average reward:", -res.fun)

                    full_path = os.path.join(sq_path, 'pickle_file')
                    os.makedirs(full_path, exist_ok=True)
                    filename = os.path.join(full_path, f'best_s.pkl')
                    with open(filename, 'wb') as f:
                            pickle.dump(res.x[0], f)
                    filename = os.path.join(full_path, f'best_Q.pkl')
                    with open(filename, 'wb') as f:
                            pickle.dump(res.x[1], f) 
                    
                    for cell in sheet[1]:  # Assumendo che i titoli siano nella prima riga
                        if cell.value == 's':
                            s_column = cell.column
                            break
                        if cell.value == 'Q':
                            Q_column = cell.column
                            break

                    if s_column is None:
                        # Creare la colonna base_stock_level se non esiste
                        s_column = sheet.max_column + 1
                        sheet.cell(row=1, column=s_column, value='s')
                    if Q_column is None:
                        # Creare la colonna base_stock_level se non esiste
                        Q_column = sheet.max_column + 1
                        sheet.cell(row=1, column=Q_column, value='Q')
                
                    
                    sheet.cell(row=i+2, column=s_column, value=np.around(res.x[0]))
                    sheet.cell(row=i+2, column=Q_column, value=np.around(res.x[1]))
                    workbook.save(excel_path)



                if brute_force_manual:
                    best_s, best_Q = train_sQ.train_sQ_policy(env, 0, 10, 0, 10, sq_path, total_timesteps)

                    for cell in sheet[1]:  # Assumendo che i titoli siano nella prima riga
                        if cell.value == 's':
                            s_column = cell.column
                            break
                        if cell.value == 'Q':
                            Q_column = cell.column
                            break

                    if s_column is None:
                        # Creare la colonna base_stock_level se non esiste
                        s_column = sheet.max_column + 1
                        sheet.cell(row=1, column=s_column, value='s')
                    if Q_column is None:
                        # Creare la colonna base_stock_level se non esiste
                        Q_column = sheet.max_column + 1
                        sheet.cell(row=1, column=Q_column, value='Q')
                
                    
                    sheet.cell(row=i+2, column=s_column, value=best_s)
                    sheet.cell(row=i+2, column=Q_column, value=best_Q)

                    workbook.save(excel_path)

            if plot_train:
                plot.plot_rewards_per_sq_level(sq_path)
            
            if eval:
                best_sq_path = os.path.join(sq_path, 'pickle_file')
                with open(best_sq_path + '/best_s.pkl', 'rb') as file:
                    best_s = pickle.load(file)
                with open(best_sq_path + '/best_Q.pkl', 'rb') as file:
                    best_Q = pickle.load(file)
                sq_class = sQ.sQpolicy(best_s, best_Q)
                eval_sQ.evaluate_policy_and_log_detailed_metrics(env, sq_class, sq_path, n_eval_episodes=20)
            if plot_eval:
                eval_sQ.save_metrics_to_dataframe(sq_path, config_details)
                plot.plot_episodes_metrics(steps, sq_path)
           
main()



